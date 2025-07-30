import streamlit as st
from dotenv import load_dotenv
import os
import gc
from openpyxl import load_workbook
from utils.scraper import extract_items, clean_html
from utils.storage import load_previous_snapshot, save_snapshot, detect_new_items, push_bulk_snapshots
from utils.fetcher import fetch_html

# --- Load environment variables ---
load_dotenv()
PASSWORD = os.getenv("APP_PASSWORD") or st.secrets["APP_PASSWORD"]
CHUNK_SIZE = 5 # Num of rows processed before pushing to GitHub

# --- Page Config & Styling ---
st.set_page_config(
    page_title="LCN Consulting - Change Detection Platform",
    page_icon="🔍",
    layout="wide",
    initial_sidebar_state="expanded"
)

st.markdown("""
<style>
    @import url('https://fonts.googleapis.com/css2?family=Inter:wght@300;400;500;600;700&display=swap');
    
    :root {
        --primary-blue: #1e3a8a;
        --secondary-blue: #3b82f6;
        --success-green: #059669;
        --error-red: #dc2626;
        --new-yellow: #fcca05;
    }
    
    .lcn-header {
        background: linear-gradient(135deg, var(--primary-blue) 0%, var(--secondary-blue) 100%);
        padding: 2rem;
        border-radius: 10px;
        color: white;
        text-align: center;
        margin-bottom: 2rem;
    }
    
    .company-name {
        font-family: 'Inter', sans-serif;
        font-size: 2.5rem;
        font-weight: 700;
        margin: 0;
        color: white;
    }
    
    .tagline {
        font-family: 'Inter', sans-serif;
        font-size: 1.1rem;
        font-weight: 400;
        color: rgba(255, 255, 255, 0.9);
        margin-top: 0.5rem;
    }
    
    .status-success {
        background-color: #ecfdf5;
        border-left: 4px solid var(--success-green);
        padding: 0.75rem;
        border-radius: 6px;
        margin: 0.5rem 0;
        color: #065f46;
    }
            
    .status-new {
        background-color: #fcf3cf;
        border-left: 4px solid var(--new-yellow);
        padding: 0.75rem;
        border-radius: 6px;
        margin: 0.5rem 0;
        color: #5f5806;
    }
    
    .status-error {
        background-color: #fef2f2;
        border-left: 4px solid var(--error-red);
        padding: 0.75rem;
        border-radius: 6px;
        margin: 0.5rem 0;
        color: #991b1b;
    }
    
    .stButton > button {
        background: linear-gradient(135deg, var(--primary-blue), var(--secondary-blue));
        color: white;
        border-radius: 8px;
        padding: 0.75rem 2rem;
        font-weight: 600;
    }
</style>
""", unsafe_allow_html=True)

# --- Header ---
def create_header():
    st.markdown("""
    <div class="lcn-header">
        <h1 class="company-name">LCN Consulting</h1>
        <p class="tagline">Detecting Competitor Website Changes</p>
    </div>
    """, unsafe_allow_html=True)

# --- Sidebar ---
def create_sidebar():
    with st.sidebar:
        st.markdown("### 📋 Platform Overview")
        st.write("Monitor competitor websites for changes.")
        
        st.markdown("### 📊 Required Data Format")
        st.write("**Excel columns needed:** `Company`, `URL`, `URL Type`")

# --- Session State for Login ---
if "authenticated" not in st.session_state:
    st.session_state.authenticated = False

# --- Login Page ---
if not st.session_state.authenticated:
    st.title("Login")
    password_input = st.text_input("Enter Password", type="password")
    if st.button("Login"):
        if password_input == PASSWORD:
            st.session_state.authenticated = True
            st.success("Login successful!")
            st.rerun()  # Immediately refresh to show upload page
        else:
            st.error("Incorrect password")
    st.stop()

# --- Main Upload Page ---
create_header()
create_sidebar()
st.markdown("## 📂 Upload Configuration File")

if "uploader_key" not in st.session_state:
    st.session_state.uploader_key = 0

if 'changes' in st.session_state:
    st.markdown("## Summary")
    st.write("### Changes")
    st.write("\n".join(st.session_state['changes']) if st.session_state['changes'] else "No changes.")
    
    st.write("### No Changes")
    st.write("\n".join(st.session_state['no_changes']) if st.session_state['no_changes'] else "None.")

    st.write("### Errors")
    st.write("\n".join(st.session_state['errors']) if st.session_state['errors'] else "No errors.")

uploaded_file = st.file_uploader(
    "Upload Competitor Configuration File",
    type=['xlsx', 'xls'],
    help="Upload Excel file containing competitor URLs",
    key=f"uploaded_file_{st.session_state.uploader_key}"
)

def excel_row_generator(file):
    try:
        wb = load_workbook(file, read_only=True, data_only=True)
    except Exception as e:
        st.error(f"Error reading Excel file: {e}")
        st.stop()
    ws = wb.active
    # Ensure required columns exist (case-insensitive)
    headers = [cell.value.strip().lower() for cell in next(ws.iter_rows(min_row=1, max_row=1))]
    required_cols = ["url", "company", "url type"]
    for col in required_cols:
        if col not in headers:
            st.error(f"Missing required column: {col}")
            st.stop()
    for row in ws.iter_rows(min_row=2, values_only=True):
        yield dict(zip(headers, row))

if uploaded_file:
    st.write(f"Processing file: {uploaded_file.name}")

    changes, no_changes, errors = [], [], []
    rows = excel_row_generator(uploaded_file)
    results_container = st.container()
    progress_bar = st.progress(0)

    buffer = []
    total_processed = 0
    total_count = sum(1 for _ in rows)
    uploaded_file.seek(0)

    # --- Process in chunks ---
    for url, company_name, url_type in rows:
        buffer.append((url, company_name, url_type))
        
        if len(buffer) >= CHUNK_SIZE:
            for u, c, t in buffer:
            
                status_box = results_container.empty()
                status_box.markdown(f"\nAccessing ({c}, {t}): {u}\n")

                try:
                    html, source, status_code = fetch_html(u)

                    if html:
                        cleaned_html = clean_html(html)
                        del html # Free memory early
                        gc.collect()
                        status_box.markdown(f"Success ({source}): {company_name}\n")

                        items, error = extract_items(cleaned_html, u)
                        if error:
                            status_box.markdown(f'<div class="status-error">🚨"⚠️ Could not extract structured content from {c} ({t}): {error}\n"</div>', unsafe_allow_html=True)
                            continue

                        previous = load_previous_snapshot(c, t)
                        new_items = detect_new_items(previous, items)
                        del previous # Free memory
                        gc.collect()

                        if new_items:
                            changes.append(f"{c} ({t})")
                            status_box.markdown(f'<div class="status-new">🆕 {c} ({t}) - Changed</div>', unsafe_allow_html=True)
                        # for item in new_items:
                        #     results_log.append(f"    - {item['title']} ({item['timestamp']})\n")
                        #     results_log.append(f"      Link: {item['link']}\n")
                        else:
                            no_changes.append(f"{c} ({t})")
                            status_box.markdown(f'<div class="status-success">✅ {c} ({t}) - No Change</div>',unsafe_allow_html=True)

                        save_snapshot(c, t, items)
                    else:
                        if status_code == 404:
                            errors.append(f"{c} ({t}) - Error {status_code}. Website does not exist.")
                            status_box.markdown(f'<div class="status-error">🚨 {c} ({t}) - Error {status_code}. Website does not exist. </div>', unsafe_allow_html=True)
                        elif status_code == 403:
                            errors.append(f"{c} ({t}) - Error {status_code}. Forbidden (bot detected).")
                            status_box.markdown(f'<div class="status-error">🚨 {c} ({t}) - Error {status_code}. Forbidden (bot detected).</div>', unsafe_allow_html=True)
                        else:
                            errors.append(f"{c} ({t}) - Error {status_code}. Failed to fetch.")
                            status_box.markdown(f'<div class="status-error">🚨 {c} ({t}) - Error {status_code}. Failed to fetch, please check URL manually.</div>', unsafe_allow_html=True)

                except Exception as error:
                    errors.append(f"{c} ({t}) - {error}")
                    status_box.markdown(f'<div class="status-error">🚨 {c} ({t}) - Error {error}</div>')
                total_processed += 1
                progress_bar.progress(total_processed / total_count)
                gc.collect()
            buffer.clear()

     # Process leftover rows
    if buffer:
        for u, c, t in buffer:

            status_box = results_container.empty()
            status_box.markdown(f"\nAccessing ({c}, {t}): {u}\n")

            try:
                html, source, status_code = fetch_html(u)

                if html:
                    cleaned_html = clean_html(html)
                    del html # Free memory early
                    gc.collect()
                    status_box.markdown(f"Success ({source}): {company_name}\n")

                    items, error = extract_items(cleaned_html, u)
                    if error:
                        status_box.markdown(f'<div class="status-error">🚨"⚠️ Could not extract structured content from {c} ({t}): {error}\n"</div>', unsafe_allow_html=True)
                        continue

                    previous = load_previous_snapshot(c, t)
                    new_items = detect_new_items(previous, items)
                    del previous # Free memory
                    gc.collect()

                    if new_items:
                        changes.append(f"{c} ({t})")
                        status_box.markdown(f'<div class="status-new">🆕 {c} ({t}) - Changed</div>', unsafe_allow_html=True)
                    else:
                        no_changes.append(f"{c} ({t})")
                        status_box.markdown(f'<div class="status-success">✅ {c} ({t}) - No Change</div>',unsafe_allow_html=True)

                    save_snapshot(c, t, items)
                else:
                    if status_code == 404:
                        errors.append(f"{c} ({t}) - Error {status_code}. Website does not exist.")
                        status_box.markdown(f'<div class="status-error">🚨 {c} ({t}) - Error {status_code}. Website does not exist. </div>', unsafe_allow_html=True)
                    elif status_code == 403:
                        errors.append(f"{c} ({t}) - Error {status_code}. Forbidden (bot detected).")
                        status_box.markdown(f'<div class="status-error">🚨 {c} ({t}) - Error {status_code}. Forbidden (bot detected).</div>', unsafe_allow_html=True)
                    else:
                        errors.append(f"{c} ({t}) - Error {status_code}. Failed to fetch.")
                        status_box.markdown(f'<div class="status-error">🚨 {c} ({t}) - Error {status_code}. Failed to fetch, please check URL manually.</div>', unsafe_allow_html=True)

            except Exception as error:
                errors.append(f"{c} ({t}) - {error}")
                status_box.markdown(f'<div class="status-error">🚨 {c} ({t}) - Error {error}</div>')
            total_processed += 1
            progress_bar.progress(total_processed / total_count)
            gc.collect()

    push_bulk_snapshots()
    st.session_state['changes'] = changes
    st.session_state['no_changes'] = no_changes
    st.session_state['errors'] = errors

    # --- Reset uploader but keep results ---
    st.session_state.uploader_key += 1
    st.rerun()

# --- Logout Button ---
if st.session_state.authenticated:
    if st.button("Logout"):
        st.session_state.authenticated = False
        st.rerun()
